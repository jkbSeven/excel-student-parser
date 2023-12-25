import openpyxl as xl

class Formatting:
    def __init__(self, inputFile: str, inplace: bool = False):
        self.inputFile = inputFile
        self.wb = xl.load_workbook(self.inputFile)
        if inplace:
            self.outputFile = inputFile
        else:
            self.outputFile = inputFile.rstrip(".xlsx") + '_formatted.xlsx'

        self.sheet = self.wb.active

    def setSheet(self, sheetName: str) -> None:
        self.sheet = self.wb[sheetName]

    def getCell(self, rowIndex: int, columnIndex: int) -> xl.cell.cell.Cell:
        return self.sheet.cell(row=rowIndex, column=columnIndex)

    @property
    def headers(self) -> list[xl.cell.cell.Cell]:
        return [cell for cell in self.sheet[1]]

    @staticmethod
    def clearCell(cell: xl.cell.cell.Cell) -> None:
        cell.font = xl.styles.Font()
        cell.border = xl.styles.Border()
        cell.fill = xl.styles.PatternFill()
        cell.number_format = xl.styles.numbers.FORMAT_GENERAL
        cell.protection = xl.styles.Protection()
        cell.alignment = xl.styles.Alignment(horizontal='left')

    def clear(self) -> None:
        for row in self.sheet:
            for cell in row:
                self.clearCell(cell)

    @staticmethod
    def bold(cell: xl.cell.cell.Cell) -> None:
        cell.font = cell.font.copy(bold=True)

    @staticmethod
    def unbold(cell: xl.cell.cell.Cell) -> None:
        cell.font = cell.font.copy(bold=False)

    @staticmethod
    def align(cell: xl.cell.cell.Cell, alignment: str = "left") -> None:
        if alignment in ["left", "center", "right"]:
            cell.alignment = cell.alignment.copy(horizontal=alignment)

    @staticmethod
    def titleValue(cell: xl.cell.cell.Cell) -> None:
        if isinstance(cell.value, str):
            cell.value = cell.value.title()

    def save(self) -> None:
        self.wb.save(self.outputFile)

    def close(self) -> None:
        self.wb.close()

