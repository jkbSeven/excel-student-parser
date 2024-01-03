import openpyxl as xl

class XlHandler:
    def __init__(self, inputFile: str, inplace: bool = False):
        self.inputFile = inputFile
        self.wb = xl.load_workbook(self.inputFile)
        self.sheet = self.wb.active

        if inplace:
            self.outputFile = inputFile
        else:
            self.outputFile = inputFile.rstrip(".xlsx") + '_out.xlsx'

    def setSheet(self, sheetName: str) -> None:
        self.sheet = self.wb[sheetName]

    def printSheet(self, min_row: int = 0, min_col: int = 0, max_row: int = None, max_col: int = None) -> None:
        if max_row is None:
            max_row = self.sheet.max_row
        if max_col is None:
            max_col = self.sheet.max_column

        for row in self.sheet.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
            for cell in row:
                if cell.value is not None:
                    print(f"{cell.value: ^20}", end="")
                else:
                    print(f"{'': ^20}", end="")
            print()

    def getCell(self, rowIndex: int, columnIndex: int) -> xl.cell.cell.Cell:
        return self.sheet.cell(row=rowIndex, column=columnIndex)

    def setValue(self, cell, value: str = None) -> None:
        if value is not None:
            cell.value = str(value)
        else:
            cell.value = None

    def findValue(self, value: str, count: int = 0) -> list[xl.cell.cell.Cell] | None:
        # count == 0: return all cells found
        if not isinstance(count, int):
            return None
        if count != 0 and count < 1:
            return None

        cells = []
        found = 0
        for row in self.sheet.iter_rows():
            for cell in row:
                if cell.value == value:
                    cells.append(cell)
                    found += 1
                    if count != 0 and found >= count:
                        return cells
                        
        return cells

    def getFirstFromColumn(self, columnIndex: int) -> tuple[int, str] | None:
        for rowIndex in range(1, self.sheet.max_row + 1):
            if (value:=self.sheet.cell(row=rowIndex, column=columnIndex).value) is not None:
                return rowIndex, value
        return None

    def getFirstFromRow(self, rowIndex: int) -> tuple[int, str] | None:
        for columnIndex in range(1, self.sheet.max_column + 1):
            if (value:=self.sheet.cell(row=rowIndex, column=columnIndex).value) is not None:
                return columnIndex, value
        return None

    @property
    def headers(self) -> list[xl.cell.cell.Cell]:
        return [cell for cell in self.sheet[1]]

    def split(self, columnIndex: int, pattern: str) -> None:
        firstFromColumn = self.getFirstFromColumn(columnIndex)
        if firstFromColumn is None:
            return

        rowIndex, value = firstFromColumn
        newColumns = len(value.strip(pattern).split(pattern))

        if newColumns < 2:
            return

        self.sheet.insert_cols(columnIndex + 1, newColumns - 1)

        for row in self.sheet.iter_rows(min_col=columnIndex, max_col=columnIndex):
            cell = row[0]
            cellSplit = cell.value.strip(pattern).split(pattern)
            while len(cellSplit) < newColumns:
                cellSplit.append(None)

            for offset in range(newColumns):
                self.sheet.cell(row=cell.row, column=(columnIndex + offset)).value = cellSplit[offset]

    def shiftColumns(self, columnIndex: int, offset: int) -> None:
        # offset < 0: shift left, offset > 0: shift right
        self.sheet.move_range(f"{xl.utils.get_column_letter(columnIndex)}1:{xl.utils.get_column_letter(self.sheet.max_column)}{self.sheet.max_row}", cols=offset)

    def strip(self, cell, pattern: str = "") -> None:
        if isinstance(cell.value, str):
            cell.value = cell.value.strip(pattern)

    def lstrip(self, cell, pattern: str = " ") -> None:
        if isinstance(cell.value, str):
            cell.value = cell.value.lstrip(pattern)

    def rstrip(self, cell, pattern: str = " ") -> None:
        if isinstance(cell.value, str):
            cell.value = cell.value.rstrip(pattern)

    def replace(self, cell, old: str, new: str) -> None:
        if isinstance(cell.value, str):
            cell.value = cell.value.replace(old, new)

    def save(self):
        self.wb.save(self.outputFile)

    def close(self):
        self.wb.close()
